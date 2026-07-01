#!/usr/bin/env python3
"""
Converts submission.csv (produced by rank.py) into the .xlsx format the
Redrob Hack2Skill portal requires for the "ranked output file" upload.

This is a separate, standalone conversion step — rank.py's reproduce_command
and the CSV output stay untouched (the CSV is still what
validate_submission.py checks against the spec). Run this only when preparing
the portal upload.

Usage:
    python convert_to_xlsx.py submission.csv submission.xlsx
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def convert(csv_path: str, xlsx_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "submission"

    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
            if c_idx == 3 and r_idx > 1:  # score column -> numeric
                try:
                    cell.value = float(val)
                except ValueError:
                    pass
            if c_idx == 2 and r_idx > 1:  # rank column -> integer
                try:
                    cell.value = int(val)
                except ValueError:
                    pass

    # reasonable column widths
    widths = {1: 18, 2: 8, 3: 10, 4: 90}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path} ({len(rows)-1} data rows)")


def main():
    if len(sys.argv) != 3:
        print("Usage: python convert_to_xlsx.py <submission.csv> <submission.xlsx>")
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
